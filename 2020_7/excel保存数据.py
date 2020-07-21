import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import csv
#CSV 也是按照一定规范书写的文本。xlsx 格式的文件是二进制的，只能被 Excel 打开。而 csv 格式的文件是纯文本，你甚至可以用记事本打开它。当用 Excel 打开 csv 文件时，会将其解析成表格形式展示.
#和 xlsx 文件相比，csv 文件占用空间和内容小，打开的速度也更快。但 csv 文件功能受限，不能存储图表、公式、图片等。

'''
wb = Workbook()
sheet = wb.active
sheet.title = '豆瓣图书Top250'
# 写入表头
header = ['书名', '评分', '链接']
sheet.append(header)

headers = { 'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.132 Safari/537.36' }
res = requests.get('https://book.douban.com/top250', headers=headers)
soup = BeautifulSoup(res.text, 'html.parser')
items = soup.find_all(class_='item')
for i in items:
  tag = i.find(class_='pl2').find('a')
  rating = i.find(class_='rating_nums').text
  name = tag['title']
  link = tag['href']
  # 写入一行数据
  row = [name, rating, link]
  sheet.append(row)
  print(name, rating, link)

# 保存为 豆瓣.xlsx
wb.save('豆瓣.xlsx')
'''
#----------------------------------------------------------
'''
from openpyxl import load_workbook

# 打开 Excel 文件
wb = load_workbook('考勤统计.xlsx')
# 选中考勤统计表这张工作表
sheet = wb['考勤统计表']
# 打印出所有工作表名称
print(wb.sheetnames)

# 打印出 A1 单元格的值
print(sheet['A1'].value)

# 打印所有单元格的值
for row in sheet.rows:
  for cell in row:
    print(cell.value)
'''
'''
#注意：open() 函数中的第二个参数 newline='' 是为了让文件内容中的换行符能被正确解析，建议在用 csv 处理文件时都加上这个参数。
with open('豆瓣.csv', 'w', newline='',encoding='utf-8') as file:
  csv_writer = csv.writer(file)
  # 写入表头
  header = ['书名', '评分', '链接']
  csv_writer.writerow(header)

  headers = {'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.132 Safari/537.36'}
  res = requests.get('https://book.douban.com/top250', headers=headers)
  soup = BeautifulSoup(res.text, 'html.parser')
  items = soup.find_all(class_='item')
  for i in items:
    tag = i.find(class_='pl2').find('a')
    rating = i.find(class_='rating_nums').text
    name = tag['title']
    link = tag['href']
    # 写入一行数据
    row = [name, rating, link]
    csv_writer.writerow(row)
    print(name, rating, link)
'''

'''
#读取csv
with open('考勤统计表.csv', newline='') as file:
  csv_reader = csv.reader(file)
  for row in csv_reader:
    print(row)
'''
