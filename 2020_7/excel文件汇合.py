from openpyxl import Workbook, load_workbook

filenames = ['2019-12-%02d-销售数据.xlsx' % (i + 1) for i in range(31)]

total_wb = Workbook()
# 选择默认的工作表
total_sheet = total_wb.active
# 给工作表重命名
total_sheet.title = '销量数据'

for filename in filenames:
  wb = load_workbook(filename)
  sheet = wb.active
  for index, row in enumerate(sheet.values):
    if filename != filenames[0] and index == 0:  # 跳过非第一个文件的表头
      continue
    total_sheet.append(row)

total_wb.save('12月销售数据汇总.xlsx')

clac_wb = Workbook()
clac_sheet = clac_wb.active
clac_sheet.title = '销量数据'

wb = load_workbook('12月销售数据汇总.xlsx')
sheet = wb.active

for index, row in enumerate(sheet.values):
  if index == 0:  # 第一个是表头
    clac_sheet.append(row + ('购买转化率', '客单价'))  # 添加两个新表头
  else:
    visitors = int(row[2])  # 访客数
    buyers = int(row[3])  # 买家数
    gmv = int(row[4])  # 交易额
    sale_rate = buyers / visitors if visitors else 0  # 购买转化率
    pct = gmv / buyers if buyers else 0  # 客单价
    clac_sheet.append(row + (sale_rate, pct))  # 添加购买转化率和客单价

clac_wb.save('12月销售计算数据汇总.xlsx')